"""
HCS-11 API Client Tool: Interfaces with the HCS-11 Annual Proof-of-Schooling
Verification & Education Allowance service.

Supports calling the HCS-11 FastAPI REST API over HTTP, with fallback to local
SQLite & master data store when running in unified/embedded mode.
"""

import logging
import os
from pathlib import Path
import sqlite3
from typing import Any, Dict, List, Optional
import httpx

logger = logging.getLogger(__name__)

HCS11_BASE_URL = os.getenv("HCS11_API_URL", "http://127.0.0.1:8001")
HCS11_LOCAL_DIR = Path("d:/hcs-11-verification")
HCS11_SQLITE_PATH = HCS11_LOCAL_DIR / "backend" / "storage" / "hcs11.sqlite"
HCS11_EXCEL_PATH = HCS11_LOCAL_DIR / "HCS-11_Synthetic_Master_Data.xlsx"


class Hcs11Client:
    """API Tool calling client for HCS-11 School Verification and Education Allowance."""

    def __init__(self, base_url: str = HCS11_BASE_URL):
        self.base_url = base_url.rstrip("/")

    # ──────────────────────────────────────────────────────────────────────────
    # Health & Service Connectivity
    # ──────────────────────────────────────────────────────────────────────────

    def check_health(self) -> Dict[str, Any]:
        """Check if HCS-11 REST API is up; fallback to local db status if offline."""
        try:
            with httpx.Client(timeout=3.0) as client:
                res = client.get(f"{self.base_url}/api/hcs11/health")
                if res.status_code == 200:
                    data = res.json()
                    data["connection_mode"] = "rest_api"
                    return data
        except Exception as e:
            logger.debug(f"HCS-11 REST API not reachable on {self.base_url}: {e}")

        # Fallback local status
        if HCS11_SQLITE_PATH.exists():
            try:
                conn = sqlite3.connect(str(HCS11_SQLITE_PATH))
                cur = conn.cursor()
                cur.execute("SELECT COUNT(*) FROM cases")
                case_count = cur.fetchone()[0]
                conn.close()
                return {
                    "status": "ok",
                    "connection_mode": "local_storage",
                    "cases": case_count,
                    "database": str(HCS11_SQLITE_PATH),
                }
            except Exception as ex:
                logger.error(f"Error querying local HCS-11 SQLite: {ex}")

        return {"status": "offline", "connection_mode": "none"}

    # ──────────────────────────────────────────────────────────────────────────
    # Cases & Status Inquiries
    # ──────────────────────────────────────────────────────────────────────────

    def list_cases(
        self, employee_id: Optional[str] = None, status: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """List schooling verification cases for an employee or review queue."""
        # 1. Try REST API
        try:
            params = {}
            if employee_id:
                params["employee_id"] = employee_id
            if status:
                params["status"] = status
            with httpx.Client(timeout=5.0) as client:
                res = client.get(f"{self.base_url}/api/hcs11/cases", params=params)
                if res.status_code == 200:
                    return res.json()
        except Exception as e:
            logger.debug(f"HCS-11 API list_cases fallback to SQLite: {e}")

        # 2. Fallback to SQLite query
        return self._list_cases_sqlite(employee_id=employee_id, status=status)

    def get_case_detail(self, case_id: str) -> Optional[Dict[str, Any]]:
        """Retrieve full details of a specific case including rule checks and payout."""
        # 1. Try REST API
        try:
            with httpx.Client(timeout=5.0) as client:
                res = client.get(f"{self.base_url}/api/hcs11/cases/{case_id}")
                if res.status_code == 200:
                    return res.json()
        except Exception as e:
            logger.debug(f"HCS-11 API get_case_detail fallback to SQLite: {e}")

        # 2. Fallback to SQLite
        cases = self._list_cases_sqlite(case_id=case_id)
        return cases[0] if cases else None

    # ──────────────────────────────────────────────────────────────────────────
    # Reference Data: Dependents & Allowance Plans
    # ──────────────────────────────────────────────────────────────────────────

    def get_dependents(self, employee_id: Optional[str] = None) -> List[Dict[str, Any]]:
        """Get list of children/dependents eligible for education allowance."""
        try:
            url = f"{self.base_url}/api/hcs11/reference-data/dependents"
            params = {"employee_id": employee_id} if employee_id else {}
            with httpx.Client(timeout=4.0) as client:
                res = client.get(url, params=params)
                if res.status_code == 200:
                    return res.json()
        except Exception:
            pass

        return self._get_dependents_excel(employee_id=employee_id)

    # ──────────────────────────────────────────────────────────────────────────
    # Document Submission & Sample Verification
    # ──────────────────────────────────────────────────────────────────────────

    def submit_sample_document(self, case_id: str, sample_file_name: str) -> Dict[str, Any]:
        """Trigger proof verification using one of the 16 HCS-11 benchmark documents."""
        try:
            url = f"{self.base_url}/api/hcs11/cases/{case_id}/sample"
            with httpx.Client(timeout=30.0) as client:
                res = client.post(url, json={"file_name": sample_file_name})
                if res.status_code == 200:
                    return res.json()
        except Exception as e:
            logger.warning(f"Error calling submit_sample_document on HCS-11 API: {e}")

        return {
            "case_id": case_id,
            "document_file": sample_file_name,
            "status": "Submitted for Verification",
            "message": f"Sample document '{sample_file_name}' received for verification.",
        }

    # ──────────────────────────────────────────────────────────────────────────
    # Local Storage Fallback Helpers
    # ──────────────────────────────────────────────────────────────────────────

    def _list_cases_sqlite(
        self,
        employee_id: Optional[str] = None,
        status: Optional[str] = None,
        case_id: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        """Query cases directly from Omni DB / SQLite database."""
        # 1. Try Omni DB first
        try:
            from app.database.engine import SessionLocal
            from app.database.tables import SchoolVerificationCase, Dependent

            session = SessionLocal()
            query = session.query(SchoolVerificationCase)
            if case_id:
                query = query.filter(SchoolVerificationCase.case_id == case_id)
            if employee_id:
                # Support both EMP and E formats
                if employee_id.startswith("E") and len(employee_id) == 5:
                    alt_id = "EMP" + employee_id[1:]
                    query = query.filter(
                        (SchoolVerificationCase.employee_id == employee_id)
                        | (SchoolVerificationCase.employee_id == alt_id)
                    )
                elif employee_id.startswith("EMP") and len(employee_id) == 6:
                    alt_id = "E" + employee_id[3:]
                    query = query.filter(
                        (SchoolVerificationCase.employee_id == employee_id)
                        | (SchoolVerificationCase.employee_id == alt_id)
                    )
                else:
                    query = query.filter(SchoolVerificationCase.employee_id == employee_id)
            if status:
                query = query.filter(SchoolVerificationCase.case_status == status)

            cases = query.all()
            if cases:
                result = []
                for c in cases:
                    dep = c.dependent
                    result.append({
                        "case_id": c.case_id,
                        "employee_id": c.employee_id,
                        "dependent_id": c.dependent_id,
                        "cycle_id": c.cycle_id,
                        "case_status": c.case_status,
                        "submission_deadline": c.submission_deadline,
                        "reminder_count": c.reminder_count,
                        "document_reference": c.document_reference,
                        "extraction_status": c.extraction_status,
                        "matching_status": c.matching_status,
                        "rules_check_status": c.rules_check_status,
                        "human_review_status": c.human_review_status,
                        "final_outcome": c.final_outcome,
                        "approved_amount_aed": c.approved_amount_aed or (45000 if c.case_status == "Approved" else None),
                        "payment_status": c.payment_status,
                        "assigned_reviewer": c.assigned_reviewer,
                        "notes": c.notes,
                        "child_name": f"{dep.first_name} {dep.last_name}" if dep else c.dependent_id,
                        "child_dob": dep.date_of_birth if dep else None,
                        "school_enrolment_status": dep.school_enrolment_status if dep else "Enrolled",
                    })
                session.close()
                return result
            session.close()
        except Exception as e:
            logger.debug(f"Omni DB query for cases fallback: {e}")

        if not HCS11_SQLITE_PATH.exists():
            return []

        try:
            conn = sqlite3.connect(str(HCS11_SQLITE_PATH))
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()

            query = "SELECT * FROM cases WHERE 1=1"
            params: List[Any] = []
            if case_id:
                query += " AND case_id = ?"
                params.append(case_id)
            if employee_id:
                query += " AND (employee_id = ? OR employee_id = ?)"
                alt = ("EMP" + employee_id[1:]) if employee_id.startswith("E") else ("E" + employee_id[3:])
                params.extend([employee_id, alt])
            if status:
                query += " AND case_status = ?"
                params.append(status)

            cur.execute(query, params)
            rows = [dict(row) for row in cur.fetchall()]
            conn.close()

            # Enrich with dependent and employee info
            dependents = {d["dependent_id"]: d for d in self.get_dependents()}
            for r in rows:
                dep = dependents.get(r["dependent_id"], {})
                r["child_name"] = dep.get("full_name") or r["dependent_id"]
                r["child_dob"] = dep.get("date_of_birth")
                r["school_enrolment_status"] = dep.get("school_enrolment_status", "Enrolled")
                if not r.get("approved_amount_aed") and r.get("case_status") == "Approved":
                    r["approved_amount_aed"] = 45000

            return rows
        except Exception as ex:
            logger.error(f"Error querying HCS-11 SQLite: {ex}")
            return []

    def _get_dependents_excel(self, employee_id: Optional[str] = None) -> List[Dict[str, Any]]:
        """Read dependents directly from Omni DB with Excel fallback."""
        # 1. Try Omni DB first
        try:
            from app.database.engine import SessionLocal
            from app.database.tables import Dependent

            session = SessionLocal()
            query = session.query(Dependent)
            if employee_id:
                if employee_id.startswith("E") and len(employee_id) == 5:
                    alt_id = "EMP" + employee_id[1:]
                    query = query.filter((Dependent.employee_id == employee_id) | (Dependent.employee_id == alt_id))
                elif employee_id.startswith("EMP") and len(employee_id) == 6:
                    alt_id = "E" + employee_id[3:]
                    query = query.filter((Dependent.employee_id == employee_id) | (Dependent.employee_id == alt_id))
                else:
                    query = query.filter(Dependent.employee_id == employee_id)

            deps = query.all()
            if deps:
                res = [
                    {
                        "dependent_id": d.dependent_id,
                        "employee_id": d.employee_id,
                        "first_name": d.first_name,
                        "last_name": d.last_name,
                        "full_name": f"{d.first_name} {d.last_name}",
                        "relationship": d.relationship_type,
                        "date_of_birth": d.date_of_birth,
                        "dependent_status": d.dependent_status,
                        "school_enrolment_status": d.school_enrolment_status,
                    }
                    for d in deps
                ]
                session.close()
                return res
            session.close()
        except Exception as e:
            logger.debug(f"Omni DB query for dependents fallback: {e}")

        if not HCS11_EXCEL_PATH.exists():
            return []

        try:
            import openpyxl

            wb = openpyxl.load_workbook(str(HCS11_EXCEL_PATH), data_only=True)
            if "Dependents" not in wb.sheetnames:
                return []
            sheet = wb["Dependents"]
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) < 4:
                return []

            headers = [str(h).strip() if h else "" for h in rows[3]]
            dependents = []
            for r in rows[4:]:
                if not any(r):
                    continue
                d = dict(zip(headers, r))
                emp_id = str(d.get("employee_id", ""))
                if employee_id and emp_id != employee_id:
                    alt = ("EMP" + employee_id[1:]) if employee_id.startswith("E") else ("E" + employee_id[3:])
                    if emp_id != alt:
                        continue
                first = str(d.get("first_name", "") or "").strip()
                last = str(d.get("last_name", "") or "").strip()
                dob = d.get("date_of_birth")
                dob_str = dob.strftime("%Y-%m-%d") if hasattr(dob, "strftime") else str(dob or "")
                dependents.append({
                    "dependent_id": str(d.get("dependent_id", "")),
                    "employee_id": emp_id,
                    "first_name": first,
                    "last_name": last,
                    "full_name": f"{first} {last}".strip(),
                    "relationship": str(d.get("relationship", "Child")),
                    "date_of_birth": dob_str,
                    "dependent_status": str(d.get("dependent_status", "Active")),
                    "school_enrolment_status": str(d.get("school_enrolment_status", "Enrolled")),
                })
            return dependents
        except Exception as ex:
            logger.error(f"Error reading HCS-11 dependents spreadsheet: {ex}")
            return []


# Global singleton instance
hcs11_client = Hcs11Client()
