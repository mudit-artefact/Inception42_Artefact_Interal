import openpyxl

wb = openpyxl.load_workbook(r"d:\HCS_01\test_scenarios_md.xlsx")
ws = wb["Unified Master QA Tracker"]

header = [c.value for c in ws[1]]
print("Headers:", header)

with open("d:/HCS_01/Backend/scripts/eval_review_dump.txt", "w", encoding="utf-8") as f:
    for row_idx in range(2, ws.max_row + 1):
        test_id = ws.cell(row=row_idx, column=2).value
        aspect = ws.cell(row=row_idx, column=3).value
        category = ws.cell(row=row_idx, column=4).value
        definition = ws.cell(row=row_idx, column=5).value
        prompts = ws.cell(row=row_idx, column=6).value
        criteria = ws.cell(row=row_idx, column=7).value
        result = ws.cell(row=row_idx, column=8).value
        status = ws.cell(row=row_idx, column=9).value
        
        f.write(f"=== ROW {row_idx} [{test_id}] ===\n")
        f.write(f"Aspect: {aspect} | Category: {category}\n")
        f.write(f"Definition: {definition}\n")
        f.write(f"Prompts: {prompts}\n")
        f.write(f"Criteria: {criteria}\n")
        f.write(f"Current Status: {status}\n")
        f.write(f"Chatbot Result:\n{result}\n\n")

print("Dumped 38 rows to eval_review_dump.txt")
