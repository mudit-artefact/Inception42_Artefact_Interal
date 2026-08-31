"""
Backend/scripts/run_and_fill_excel.py
Executes all 38 test scenarios from test_scenarios_md.xlsx against the HCS-01 Concierge backend,
collects the responses (joining multiple queries with commas), and writes them directly into
the 'Chatbot Result' column of test_scenarios_md.xlsx using openpyxl.
"""

import os
import re
import sys
import time
import uuid
import openpyxl
from pathlib import Path

# Fix Windows console encoding for logging
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# Add backend directory to sys.path
BACKEND_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BACKEND_DIR))

from fastapi.testclient import TestClient
from app.main import app

EXCEL_FILE = Path(r"d:\HCS_01\test_scenarios_md.xlsx")


def extract_prompts(test_id: str, text: str) -> list[str]:
    """Extract individual query strings from column F text."""
    if not isinstance(text, str) or not text.strip():
        return []
    
    text = text.strip()
    
    # Check for T1: ... T2: ... T3: ... (Multi-turn dialogue)
    if re.search(r'T\d+:', text):
        pattern = r"T\d+:\s*(?:\([^)]*\)\s*)?['\"]?([^'\"\n]+)['\"]?"
        matches = re.findall(pattern, text)
        if matches:
            return [m.strip().strip("'\"") for m in matches if m.strip()]
    
    # Check for arrow sequences (A → B → C)
    if '→' in text or '->' in text:
        parts = re.split(r'\s*(?:->|→)\s*', text)
        return [p.strip().strip("'\"“”`") for p in parts if p.strip()]
        
    # Check for multi-line queries e.g. "Q1"\n"Q2"
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    if len(lines) > 1:
        extracted = []
        for line in lines:
            clean = line.strip().strip("'\"“”`")
            if clean:
                extracted.append(clean)
        return extracted
    
    # Row 38 comma-separated list of non-HR requests
    if test_id == 'TAX-38' and ',' in text:
        return [p.strip() for p in text.split(',') if p.strip()]
        
    return [text.strip().strip("'\"“”`")]


def run_all_excel_scenarios():
    print("=" * 70)
    print(f"🚀 Starting Automated Excel Evaluation on: {EXCEL_FILE}")
    print("=" * 70)

    if not EXCEL_FILE.exists():
        print(f"❌ Error: {EXCEL_FILE} does not exist!")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Unified Master QA Tracker"]
    
    # Locate headers
    header_row = [cell.value for cell in ws[1]]
    col_emp = header_row.index("employee_id") + 1
    col_test_id = header_row.index("Test ID") + 1
    col_prompts = header_row.index("Example User Prompts") + 1
    col_result = header_row.index("Chatbot Result") + 1

    total_rows = ws.max_row - 1
    print(f"📋 Loaded {total_rows} test scenarios from 'Unified Master QA Tracker' sheet.\n")

    run_token = uuid.uuid4().hex[:6]

    with TestClient(app) as client:
        for row_idx in range(2, ws.max_row + 1):
            item_num = row_idx - 1
            test_id = str(ws.cell(row=row_idx, column=col_test_id).value or f"ROW-{item_num}").strip()
            raw_emp_id = str(ws.cell(row=row_idx, column=col_emp).value or "EMP_001").strip()
            emp_id = raw_emp_id.replace("_", "").strip() or "EMP001"
            
            raw_prompts = str(ws.cell(row=row_idx, column=col_prompts).value or "")
            queries = extract_prompts(test_id, raw_prompts)

            if not queries:
                print(f"[{item_num:2d}/{total_rows}] ⚠️ {test_id}: No prompt found in column F. Skipping.")
                continue

            # Determine if this row is a single multi-turn dialogue
            is_multi_turn = ("T1:" in raw_prompts) or ("→" in raw_prompts) or ("->" in raw_prompts)
            session_id = f"excel-eval-{test_id}-{run_token}"

            print(f"[{item_num:2d}/{total_rows}] 🔍 Running {test_id} (Emp: {emp_id}, {len(queries)} query/queries)...")
            
            responses = []

            for q_idx, query in enumerate(queries, 1):
                current_conv_id = session_id if is_multi_turn else f"{session_id}-q{q_idx}"
                
                t_start = time.time()
                try:
                    res = client.post(
                        "/api/v1/hcs01/query",
                        json={
                            "query": query,
                            "employee_id": emp_id,
                            "conversation_id": current_conv_id
                        },
                        timeout=90.0
                    )
                    elapsed = round(time.time() - t_start, 2)
                    
                    if res.status_code == 200:
                        data = res.json()
                        ans = data.get("answer", "").strip()
                        if not ans and data.get("clarification_question"):
                            ans = data.get("clarification_question").strip()
                        clean_ans = re.sub(r'\s*\n+\s*', ' ', ans)
                        responses.append(clean_ans)
                        print(f"    ├─ Q{q_idx} ({elapsed}s): {query[:45]}... -> [{clean_ans[:50]}...]")
                    else:
                        err_msg = f"[Error HTTP {res.status_code}: {res.text[:100]}]"
                        responses.append(err_msg)
                        print(f"    ├─ Q{q_idx} ({elapsed}s) ❌ FAILED: HTTP {res.status_code}")
                except Exception as e:
                    err_msg = f"[Exception: {str(e)}]"
                    responses.append(err_msg)
                    print(f"    ├─ Q{q_idx} ❌ Exception: {e}")

            # Join multiple responses with commas as requested
            final_result_str = ", ".join(responses) if responses else ""

            # Update the cell directly in openpyxl
            ws.cell(row=row_idx, column=col_result).value = final_result_str
            
            # Save workbook after every row
            wb.save(EXCEL_FILE)
            print(f"    └─ ✅ Saved Result for {test_id}\n")

    print("=" * 70)
    print(f"🎉 ALL {total_rows} TEST SCENARIOS COMPLETED AND SAVED!")
    print(f"📁 Updated Excel File: {EXCEL_FILE}")
    print("=" * 70)


if __name__ == "__main__":
    run_all_excel_scenarios()
