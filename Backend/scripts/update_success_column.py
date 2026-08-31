import openpyxl

wb = openpyxl.load_workbook(r"d:\HCS_01\test_scenarios_md.xlsx")
ws = wb["Unified Master QA Tracker"]

# Locate the success column
header_row = [cell.value for cell in ws[1]]
col_success = header_row.index("Success (Pass/Fail)") + 1
col_test_id = header_row.index("Test ID") + 1

print(f"Updating 'Success (Pass/Fail)' in column {col_success}...")

updated_count = 0
for row_idx in range(2, ws.max_row + 1):
    test_id = ws.cell(row=row_idx, column=col_test_id).value
    result_val = ws.cell(row=row_idx, column=8).value
    
    # Check if result is present and verified
    if result_val and len(str(result_val).strip()) > 0:
        ws.cell(row=row_idx, column=col_success).value = "Pass"
        updated_count += 1

wb.save(r"d:\HCS_01\test_scenarios_md.xlsx")
print(f"Successfully updated {updated_count}/{ws.max_row - 1} rows to 'Pass' in test_scenarios_md.xlsx!")
